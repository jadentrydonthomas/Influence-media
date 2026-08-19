#!/usr/bin/env python3
"""Independent verification of the dashboard's figures.

Reads the fixture workbooks directly with openpyxl, recomputes every headline
number from scratch, and compares against test/figures.json produced by
test/dump-figures.mjs. Nothing here shares code with the dashboard, so an
agreement is real evidence rather than a restatement.

    node test/dump-figures.mjs && python3 test/audit.py
"""
import collections
import datetime
import json
import os
import re
import sys
import warnings

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FIX = os.path.join(ROOT, "fixtures")
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
WEEKS = {1: "Week 1 - 2026.xlsm", 2: "Week 2 - 2026.xlsm", 3: "Week 3 - 2026.xlsm"}
ORDER_LOG = "OrderLog_1-10.xlsx"

# Column positions confirmed against the real workbooks (1-based).
Q_NO, Q_CUST, Q_SQFT, Q_SCHR, Q_ENGR, Q_EST, Q_HRS, Q_ONTIME, Q_PRICE = 2, 5, 8, 12, 13, 14, 15, 17, 19
O_JOB, O_ENTRY, O_TONS, O_QUOTE = 1, 8, 13, 26


def base_key(value):
    m = re.match(r"^([A-Z0-9]{2,4})-(\d{4,6})", str(value or "").strip().upper())
    return "%s-%s" % (m.group(1), m.group(2)) if m else ""


def monday_of(week):
    # 2026-01-05 is the first Monday of 2026.
    return datetime.date(2026, 1, 5) + datetime.timedelta(days=7 * (week - 1))


def read_quotes():
    rows = []
    for week, name in WEEKS.items():
        wb = openpyxl.load_workbook(os.path.join(FIX, name), data_only=True, read_only=True)
        for offset, day in enumerate(DAYS):
            quote_date = monday_of(week) + datetime.timedelta(days=offset)
            for r in wb[day].iter_rows(min_row=6, max_col=20, values_only=True):
                if r[Q_NO - 1] is None or not str(r[Q_NO - 1]).strip():
                    continue
                rows.append({
                    "key": base_key(r[Q_NO - 1]), "week": week, "date": quote_date,
                    "customer": r[Q_CUST - 1], "sqft": r[Q_SQFT - 1], "price": r[Q_PRICE - 1],
                    "schr": r[Q_SCHR - 1], "engr": r[Q_ENGR - 1], "est": r[Q_EST - 1],
                    "hrs": r[Q_HRS - 1], "ontime": str(r[Q_ONTIME - 1] or "").strip().upper(),
                })
        wb.close()
    return rows


def read_orders():
    orders, blank, unparsable = [], 0, []
    wb = openpyxl.load_workbook(os.path.join(FIX, ORDER_LOG), data_only=True, read_only=True)
    for r in wb["Sheet1"].iter_rows(min_row=3, max_col=27, values_only=True):
        if r[O_JOB - 1] is None or not str(r[O_JOB - 1]).strip():
            continue
        raw = str(r[O_QUOTE - 1] or "").strip()
        key = base_key(raw)
        if not raw:
            blank += 1
        elif not key:
            unparsable.append(raw)
        orders.append({"key": key, "entry": r[O_ENTRY - 1], "tons": r[O_TONS - 1]})
    wb.close()
    return orders, blank, unparsable


def collapse(rows):
    groups = collections.OrderedDict()
    for row in rows:
        groups.setdefault(row["key"], []).append(row)
    out = []
    for key, group in groups.items():
        def last(field):
            for item in reversed(group):
                v = item[field]
                if isinstance(v, (int, float)) and v:
                    return v
            return 0

        def last_ontime():
            # Verified across every fixture row: EARLY <=> Done < Due,
            # LATE <=> Done > Due, N/A <=> Done = Due. N/A is a real result
            # meaning "delivered on the due date", not a missing value.
            v = ""
            for item in group:
                label = item["ontime"]
                if label in ("EARLY", "LATE"):
                    v = label
                elif label in ("N/A", "NA"):
                    v = "ONTIME"
            return v

        out.append({
            "key": key, "week": group[0]["week"], "date": group[0]["date"],
            "customer": group[0]["customer"], "price": last("price"), "sqft": last("sqft"),
            "engr": str(group[0]["engr"] or "").strip().upper() or "(none)",
            "est": str(group[0]["est"] or "").strip().upper() or "(none)",
            "schr": str(group[0]["schr"] or "").strip().upper() or "(none)",
            "ontime": last_ontime(),
            "hrs": sum(i["hrs"] for i in group if isinstance(i["hrs"], (int, float))),
        })
    return out


def money(millions):
    return "$%.1fM" % millions


def main():
    figures_path = os.path.join(ROOT, "test", "figures.json")
    if not os.path.exists(figures_path):
        print("Run: node test/dump-figures.mjs first")
        return 2
    shown = json.load(open(figures_path))

    rows = read_quotes()
    opps = collapse(rows)
    orders, blank, unparsable = read_orders()
    order_keys = collections.Counter(o["key"] for o in orders if o["key"])
    dated = [o["entry"] for o in orders if isinstance(o["entry"], datetime.datetime)]
    cutoff = max(dated).date()

    for o in opps:
        o["won"] = o["key"] in order_keys
        o["exposure"] = (cutoff - o["date"]).days

    checks, failures = [], 0

    def check(name, expected, actual):
        nonlocal failures
        ok = str(expected) == str(actual)
        if not ok:
            failures += 1
        checks.append((ok, name, expected, actual))

    print("Recomputed from the workbooks with openpyxl, independent of the app.\n")
    print("  quote rows              %d" % len(rows))
    print("  distinct opportunities  %d" % len(opps))
    print("  order rows              %d  (blank Quote# %d, unparsable %d: %s)"
          % (len(orders), blank, len(unparsable), ", ".join(unparsable) or "none"))
    print("  order-entry cutoff      %s\n" % cutoff)

    for scope, minimum in (("all", None), ("mature", 30), ("high", 50)):
        cohort = [o for o in opps if minimum is None or o["exposure"] >= minimum]
        wins = [o for o in cohort if o["won"]]
        value = sum(o["price"] for o in cohort)
        scored = [o for o in cohort if o["ontime"] in ("EARLY", "LATE", "ONTIME")]
        early = [o for o in scored if o["ontime"] in ("EARLY", "ONTIME")]
        s = shown["scopes"][scope]
        label = {"all": "All quotes", "mature": "30+ days", "high": "50+ days"}[scope]
        check("%s · opportunities" % label, len(cohort), s["quotes"])
        check("%s · quote wins" % label, len(wins), s["wins"])
        check("%s · conversion" % label, "%.1f%%" % (len(wins) / len(cohort) * 100), s["conv"])
        check("%s · quoted value" % label, money(value / 1e6), s["value"])
        check("%s · unconverted" % label, len(cohort) - len(wins), s["unconverted"])
        if scored:
            check("%s · met the due date" % label, "%.1f%%" % (len(early) / len(scored) * 100), s["onTime"])
            check("%s · on-time denominator" % label,
                  "%d/%d scored" % (len(scored), len(cohort)), s["onTimeCoverage"])

    cohort = [o for o in opps if o["exposure"] >= 30]
    wins = [o for o in cohort if o["won"]]
    check("booked jobs", sum(order_keys[o["key"]] for o in wins),
          int(re.search(r"(\d+) jobs booked", " ".join(shown["scopes"]["mature"]["kpis"])).group(1)))

    # Value bands, as shown by the Value bands lens.
    bands = [("Under $250k", 0, 250e3), ("$250k – $1M", 250e3, 1e6),
             ("$1M – $5M", 1e6, 5e6), ("$5M and above", 5e6, float("inf"))]
    lens = shown["lenses"]["bands"]
    for label, lo, hi in bands:
        members = [o for o in cohort if o["price"] and lo <= o["price"] < hi]
        if not members:
            continue
        band_wins = [o for o in members if o["won"]]
        rate = "%.1f%%" % (len(band_wins) / len(members) * 100)
        pattern = re.escape(label) + r"\s+" + re.escape(rate) + r"\s+" + str(len(members)) + r" quotes"
        check("band %s" % label, "shown as %s on %d quotes" % (rate, len(members)),
              "shown as %s on %d quotes" % (rate, len(members)) if re.search(pattern, lens) else "NOT FOUND in lens text")

    # People: quote-engineer volumes must reconcile to the cohort.
    eng_counts = collections.Counter(o["engr"] for o in cohort)
    listed = shown["people"]["engineers"]
    total_listed = 0
    for row in listed:
        # The roster row prints "41 quotes · 2 wins · 2 jobs"; it used to read
        # "41 opportunities". Accept either so a copy edit cannot silently
        # zero this check.
        m = re.search(r"(\d+) (?:opportunities|quotes)\b", row)
        if m:
            total_listed += int(m.group(1))
    check("quote-engineer volumes sum to the cohort", len(cohort), total_listed)
    check("distinct engineer codes present", len(eng_counts) <= len(listed), True)

    print("Comparison against what the dashboard displayed:\n")
    for ok, name, expected, actual in checks:
        mark = "ok  " if ok else "FAIL"
        print("  %s  %-42s recomputed=%-16s shown=%s" % (mark, name, expected, actual))
    print("\n%d/%d agree" % (len(checks) - failures, len(checks)))
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
